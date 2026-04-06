import io
import os
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from models import db, User, EfficacyCatalog, Request, Message, MessageRead, Notification, RawDataRequest

load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "bt-lims-secret-key-2026")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///lims.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "로그인이 필요합니다."


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.template_filter('week_label')
def week_label_filter(value):
    """Convert '2026-03-W2' to '3월 둘째주'"""
    if not value:
        return '-'
    try:
        parts = value.split('-')
        month = int(parts[1])
        week = int(parts[2].replace('W', ''))
        week_names = ['첫째주', '둘째주', '셋째주', '넷째주', '다섯째주']
        return f"{month}월 {week_names[week-1]}"
    except (IndexError, ValueError):
        return value


# ── Helper: sync parent request status from children ─
def sync_parent_status(parent):
    """Update parent request status based on children's statuses."""
    if not parent or parent.status == "submitted":
        return
    children = parent.children
    if not children:
        return

    statuses = [c.status for c in children]

    if all(s == "completed" for s in statuses):
        parent.status = "completed"
        parent.completed_at = datetime.now(timezone.utc)
    elif any(s == "documenting" for s in statuses):
        parent.status = "documenting"
    elif any(s == "in_progress" for s in statuses):
        parent.status = "in_progress"
    else:
        parent.status = "approved_parent"

    db.session.commit()


# ── Helper: unread message count ─────────────────────
def get_unread_count(request_id, user_id):
    """Get count of unread messages for a user on a request."""
    read = MessageRead.query.filter_by(request_id=request_id, user_id=user_id).first()
    last_read = read.last_read_id if read else 0
    return Message.query.filter(
        Message.request_id == request_id,
        Message.id > last_read,
        Message.sender_id != user_id
    ).count()


def mark_as_read(request_id, user_id):
    """Mark all messages as read for a user on a request."""
    last_msg = Message.query.filter_by(request_id=request_id).order_by(Message.id.desc()).first()
    if not last_msg:
        return
    read = MessageRead.query.filter_by(request_id=request_id, user_id=user_id).first()
    if read:
        read.last_read_id = last_msg.id
    else:
        db.session.add(MessageRead(request_id=request_id, user_id=user_id, last_read_id=last_msg.id))
    db.session.commit()


# ── Auth ─────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()
        if user and check_password_hash(user.password, request.form["password"]):
            login_user(user)
            return redirect(url_for("dashboard"))
        flash("아이디 또는 비밀번호가 올바르지 않습니다.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Dashboard ────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    if current_user.role == "admin":
        # Admin sees only parent/standalone requests (no children)
        base = Request.query.filter(Request.parent_id.is_(None))
        stats = {
            "submitted": base.filter_by(status="submitted").count(),
            "pending": base.filter(Request.status.in_(["pending", "approved_parent"])).count(),
            "in_progress": base.filter_by(status="in_progress").count() +
                           base.filter_by(status="approved_parent").filter(
                               Request.children.any(Request.status == "in_progress")).count(),
            "documenting": base.filter_by(status="documenting").count(),
            "completed": base.filter_by(status="completed").count(),
        }
        review_queue = base.filter_by(status="submitted").order_by(Request.created_at.desc()).all()
        recent = base.filter(Request.status.notin_(["submitted"])).order_by(Request.created_at.desc()).limit(10).all()

        # Unread counts for admin: submitted (specialized chat) + all child chats
        unread = {}
        for req in review_queue:
            if req.has_specialized:
                cnt = get_unread_count(req.id, current_user.id)
                if cnt > 0:
                    unread[req.id] = cnt
        for req in recent:
            if req.children:
                for child in req.children:
                    if child.status == "in_progress":
                        cnt = get_unread_count(child.id, current_user.id)
                        if cnt > 0:
                            unread[child.id] = cnt
            elif req.status == "in_progress":
                cnt = get_unread_count(req.id, current_user.id)
                if cnt > 0:
                    unread[req.id] = cnt

        return render_template("dashboard.html", stats=stats, recent=recent,
                               review_queue=review_queue, unread=unread)

    elif current_user.role == "researcher":
        # Show parent requests that have pending non-specialized children,
        # or standalone pending requests
        available_parents = Request.query.filter(
            Request.parent_id.is_(None),
            Request.children.any(
                db.and_(Request.status == "pending", Request.is_specialized_child == False)
            )
        ).all()
        available_standalone = Request.query.filter(
            Request.parent_id.is_(None),
            Request.status == "pending",
            ~Request.children.any()
        ).all()
        available = available_parents + available_standalone

        # Specialized requests: assigned to me OR assigned to my team
        specialized_available = Request.query.filter(
            Request.is_specialized_child == True,
            Request.status == "pending",
            db.or_(
                Request.assigned_researcher_id == current_user.id,
                Request.assigned_team == current_user.team
            )
        ).all()

        # My work: find parent requests where I have in_progress/documenting children
        my_child_ids = [r.parent_id for r in Request.query.filter_by(researcher_id=current_user.id).filter(
            Request.status.in_(["in_progress", "documenting"]), Request.parent_id.isnot(None)
        ).all() if r.parent_id]
        my_work_parents = Request.query.filter(Request.id.in_(my_child_ids)).all() if my_child_ids else []
        my_work_standalone = Request.query.filter_by(researcher_id=current_user.id).filter(
            Request.parent_id.is_(None), Request.status.in_(["in_progress", "documenting"])
        ).all()
        my_work = my_work_parents + my_work_standalone

        completed_count = Request.query.filter_by(researcher_id=current_user.id, status="completed").count()
        completed_list = Request.query.filter_by(researcher_id=current_user.id, status="completed").order_by(
            Request.completed_at.desc()).all()
        # Load chat history for completed requests
        completed_chats = {}
        for req in completed_list:
            msgs = Message.query.filter_by(request_id=req.id).order_by(Message.created_at.asc()).all()
            if msgs:
                completed_chats[req.id] = msgs
        file_requests = Request.query.filter_by(researcher_id=current_user.id, file_requested=True).all()

        # Unread counts for my child requests
        unread = {}
        for req in my_work:
            if req.children:
                for child in req.children:
                    if child.researcher_id == current_user.id and child.status == "in_progress":
                        cnt = get_unread_count(child.id, current_user.id)
                        if cnt > 0:
                            unread[child.id] = cnt
            elif req.status == "in_progress":
                cnt = get_unread_count(req.id, current_user.id)
                if cnt > 0:
                    unread[req.id] = cnt

        return render_template("dashboard_researcher.html", available=available, my_work=my_work,
                               completed_count=completed_count, completed_list=completed_list,
                               completed_chats=completed_chats,
                               file_requests=file_requests, unread=unread,
                               specialized_available=specialized_available)

    else:  # requester - only parent/standalone requests
        my_requests = Request.query.filter_by(requester_id=current_user.id).filter(
            Request.parent_id.is_(None)
        ).order_by(Request.created_at.desc()).all()

        unread = {}
        for req in my_requests:
            if req.children:
                for child in req.children:
                    if child.status == "in_progress":
                        cnt = get_unread_count(child.id, current_user.id)
                        if cnt > 0:
                            unread[child.id] = cnt
            elif req.status == "in_progress":
                cnt = get_unread_count(req.id, current_user.id)
                if cnt > 0:
                    unread[req.id] = cnt

        return render_template("dashboard_requester.html", my_requests=my_requests, unread=unread)


# ── Request Submit (Requester) ───────────────────────
@app.route("/requests/new", methods=["GET", "POST"])
@login_required
def new_request():
    if current_user.role not in ("requester", "admin", "researcher"):
        flash("의뢰서를 제출할 수 없습니다.", "error")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        has_control = request.form.get("has_control") == "on"
        urgency = request.form.get("urgency", "보통")
        deadline = request.form.get("deadline_date", "") if urgency == "긴급" else ""
        urgent_reason = request.form.get("urgent_reason", "") if urgency == "긴급" else ""

        if current_user.role == "admin":
            requester_id = int(request.form["requester_id"])
            requester = db.session.get(User, requester_id)
            req_team = requester.team if requester else ""
        else:  # requester or researcher
            requester_id = current_user.id
            req_team = current_user.team

        has_specialized = request.form.get("has_specialized") == "on"
        specialized_types = request.form.get("specialized_types", "") if has_specialized else ""
        specialized_notes = request.form.get("specialized_notes", "") if has_specialized else ""

        # Merge common + specialized efficacy types
        all_efficacy = request.form["efficacy_types"]
        if specialized_types:
            if all_efficacy:
                all_efficacy = all_efficacy + ", " + specialized_types
            else:
                all_efficacy = specialized_types

        new_req = Request(
            material_name=request.form["material_name"],
            requester_id=requester_id,
            requester_team=req_team,
            efficacy_types=all_efficacy,
            concentration=request.form["concentration"],
            sample_count=int(request.form.get("sample_count", 1)),
            characteristics=request.form.get("characteristics", ""),
            solvent=request.form.get("solvent", ""),
            has_control=has_control,
            control_name=request.form.get("control_name", "") if has_control else "",
            control_concentration=request.form.get("control_concentration", "") if has_control else "",
            urgency=urgency,
            deadline=deadline,
            urgent_reason=urgent_reason,
            sample_return=request.form.get("sample_return", ""),
            has_specialized=has_specialized,
            specialized_types=specialized_types,
            specialized_notes=specialized_notes,
            notes=request.form.get("notes", ""),
            status="submitted",
            year=datetime.now().year,
        )
        db.session.add(new_req)
        db.session.commit()
        # Notify admins
        admins = User.query.filter_by(role="admin").all()
        for admin in admins:
            create_notification(admin.id, f"새 의뢰가 접수되었습니다: [{new_req.material_name}]", f"/requests/{new_req.id}")
        flash("의뢰서가 제출되었습니다. 관리자 검토 후 실험자에게 배포됩니다.", "success")
        return redirect(url_for("dashboard"))

    requesters = User.query.filter_by(role="requester").all()
    catalog = EfficacyCatalog.query.all()
    return render_template("request_new.html", requesters=requesters, catalog=catalog)


# ── Admin Approve (submitted → split by efficacy → pending) ──
@app.route("/requests/<int:req_id>/approve", methods=["POST"])
@login_required
def approve_request(req_id):
    if current_user.role != "admin":
        flash("관리자만 승인할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req or req.status != "submitted":
        flash("승인할 수 없는 의뢰입니다.", "error")
        return redirect(url_for("dashboard"))

    efficacies = [e.strip() for e in req.efficacy_types.split(",") if e.strip()]
    has_spec = req.has_specialized and req.specialized_notes

    # Single efficacy, no specialized → simple approve
    if len(efficacies) <= 1 and not has_spec:
        req.status = "pending"
        db.session.commit()
        create_notification(req.requester_id, f"의뢰 [{req.material_name}]이 승인되었습니다.", f"/requests/{req_id}")
        flash(f"의뢰 [{req.material_name}]을 승인하여 실험자에게 배포했습니다.", "success")
        return redirect(url_for("request_detail", req_id=req_id))

    # Multi-efficacy or has specialized → split into children
    req.status = "approved_parent"
    for eff in efficacies:
        child = Request(
            parent_id=req.id,
            material_name=req.material_name,
            requester_id=req.requester_id,
            requester_team=req.requester_team,
            efficacy_types=eff,
            concentration=req.concentration,
            sample_count=req.sample_count,
            characteristics=req.characteristics,
            has_control=req.has_control,
            control_name=req.control_name,
            control_concentration=req.control_concentration,
            urgency=req.urgency,
            deadline=req.deadline,
            sample_return=req.sample_return,
            notes=req.notes,
            status="pending",
            year=req.year,
        )
        db.session.add(child)

    # If specialized, create specialized children (status=pending but not yet assigned)
    if has_spec:
        spec_child = Request(
            parent_id=req.id,
            material_name=req.material_name,
            requester_id=req.requester_id,
            requester_team=req.requester_team,
            efficacy_types="전문평가",
            concentration=req.concentration,
            sample_count=req.sample_count,
            characteristics=req.characteristics,
            urgency=req.urgency,
            deadline=req.deadline,
            sample_return=req.sample_return,
            notes=req.specialized_notes,
            status="pending",
            year=req.year,
            is_specialized_child=True,
            has_specialized=True,
            specialized_notes=req.specialized_notes,
        )
        db.session.add(spec_child)

    db.session.commit()
    flash(f"의뢰 [{req.material_name}]을 효능별로 분리하여 배포했습니다.", "success")
    return redirect(url_for("request_detail", req_id=req_id))


# ── Admin: Assign researcher to specialized request ──
@app.route("/requests/<int:req_id>/assign", methods=["POST"])
@login_required
def assign_researcher(req_id):
    if current_user.role != "admin":
        flash("관리자만 배정할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req or not req.is_specialized_child:
        flash("전문평가 의뢰만 배정할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    assign_type = request.form.get("assign_type", "individual")
    researcher_id = request.form.get("researcher_id", "")
    team = request.form.get("assign_team", "")
    memo = request.form.get("admin_memo", "")

    if assign_type == "team" and team:
        req.assigned_team = team
        req.assigned_researcher_id = None
        req.admin_memo = memo
        req.efficacy_types = request.form.get("efficacy_label", "전문평가")
        db.session.commit()
        flash(f"전문평가를 {team} 팀에 배정했습니다.", "success")
        return redirect(url_for("request_detail", req_id=req.parent_id or req_id))

    if not researcher_id:
        flash("실험자 또는 팀을 선택해주세요.", "error")
        return redirect(url_for("request_detail", req_id=req.parent_id or req_id))

    researcher = db.session.get(User, int(researcher_id))
    if not researcher or researcher.role != "researcher":
        flash("올바른 실험자를 선택해주세요.", "error")
        return redirect(url_for("request_detail", req_id=req.parent_id or req_id))

    req.assigned_researcher_id = researcher.id
    req.assigned_team = ""
    req.admin_memo = memo
    req.efficacy_types = request.form.get("efficacy_label", "전문평가")
    db.session.commit()

    flash(f"전문평가를 {researcher.name}에게 배정했습니다.", "success")
    return redirect(url_for("request_detail", req_id=req.parent_id or req_id))


@app.route("/requests/<int:req_id>/reject", methods=["POST"])
@login_required
def reject_request(req_id):
    if current_user.role != "admin":
        flash("관리자만 반려할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req or req.status != "submitted":
        flash("반려할 수 없는 의뢰입니다.", "error")
        return redirect(url_for("dashboard"))

    req.status = "rejected"
    db.session.commit()
    flash(f"의뢰 [{req.material_name}]을 반려했습니다.", "info")
    return redirect(url_for("request_list"))


@app.route("/requests")
@login_required
def request_list():
    status_filter = request.args.get("status", "all")
    year_filter = request.args.get("year", "all")
    search_query = request.args.get("q", "").strip()

    q = Request.query
    # All roles: hide child requests, show parent/standalone only
    q = q.filter(Request.parent_id.is_(None))

    if search_query:
        q = q.filter(Request.material_name.ilike(f"%{search_query}%"))
    if status_filter == "pending":
        q = q.filter(Request.status.in_(["pending", "approved_parent"]))
    elif status_filter != "all":
        q = q.filter_by(status=status_filter)
    if year_filter != "all":
        q = q.filter_by(year=int(year_filter))

    requests_list = q.order_by(Request.created_at.desc()).all()
    years = db.session.query(db.func.distinct(Request.year)).order_by(Request.year.desc()).all()
    years = [y[0] for y in years if y[0]]
    return render_template("request_list.html", requests=requests_list, status_filter=status_filter,
                           year_filter=year_filter, years=years, search_query=search_query)


@app.route("/requests/<int:req_id>")
@login_required
def request_detail(req_id):
    req = db.session.get(Request, req_id)
    if not req:
        flash("의뢰를 찾을 수 없습니다.", "error")
        return redirect(url_for("request_list"))

    messages = []
    chat_enabled = False
    # Chat: researcher-requester during in_progress (parent or child)
    if req.status == "in_progress" and req.researcher_id:
        chat_enabled = (current_user.id == req.requester_id or current_user.id == req.researcher_id)
        messages = Message.query.filter_by(request_id=req_id).order_by(Message.created_at.asc()).all()
    # Admin-requester chat during submitted (specialized)
    elif req.status == "submitted" and req.has_specialized:
        chat_enabled = (current_user.id == req.requester_id or current_user.role == "admin")
        messages = Message.query.filter_by(request_id=req_id).order_by(Message.created_at.asc()).all()

    # For parent: build per-child chat data for template
    child_chats = {}
    researchers_list = []
    if req.children:
        for child in req.children:
            if child.researcher_id and child.status in ("in_progress", "documenting", "completed"):
                if current_user.id == req.requester_id or current_user.id == child.researcher_id or current_user.role == "admin":
                    child_chats[child.id] = Message.query.filter_by(request_id=child.id).order_by(Message.created_at.asc()).all()
        # Load researchers for admin assignment
        if current_user.role == "admin":
            researchers_list = User.query.filter_by(role="researcher").all()

    # Mark messages as read
    if chat_enabled:
        mark_as_read(req_id, current_user.id)
    # Mark child chats as read
    for child_id in child_chats:
        mark_as_read(child_id, current_user.id)

    # Raw data requests for this request and its children
    raw_data_reqs = {}
    if req.status == "completed":
        if req.children:
            for child in req.children:
                rdrs = RawDataRequest.query.filter_by(request_id=child.id).all()
                if rdrs:
                    raw_data_reqs[child.id] = rdrs
        else:
            rdrs = RawDataRequest.query.filter_by(request_id=req.id).all()
            if rdrs:
                raw_data_reqs[req.id] = rdrs

    return render_template("request_detail.html", req=req, messages=messages,
                           chat_enabled=chat_enabled, child_chats=child_chats,
                           researchers=researchers_list, raw_data_reqs=raw_data_reqs)


# ── Admin Edit/Delete ────────────────────────────────
@app.route("/requests/<int:req_id>/edit", methods=["GET", "POST"])
@login_required
def edit_request(req_id):
    if current_user.role != "admin":
        flash("관리자만 수정할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req:
        flash("의뢰를 찾을 수 없습니다.", "error")
        return redirect(url_for("request_list"))

    if request.method == "POST":
        req.material_name = request.form["material_name"]
        requester_id = int(request.form["requester_id"])
        requester = db.session.get(User, requester_id)
        req.requester_id = requester_id
        req.requester_team = requester.team if requester else ""
        req.efficacy_types = request.form["efficacy_types"]
        req.concentration = request.form["concentration"]
        req.sample_count = int(request.form.get("sample_count", 1))
        req.characteristics = request.form.get("characteristics", "")
        has_control = request.form.get("has_control") == "on"
        req.has_control = has_control
        req.control_name = request.form.get("control_name", "") if has_control else ""
        req.control_concentration = request.form.get("control_concentration", "") if has_control else ""
        urgency = request.form.get("urgency", "보통")
        req.urgency = urgency
        req.deadline = request.form.get("deadline_date", "") if urgency == "기타" else ""
        req.sample_return = request.form.get("sample_return", "")
        req.notes = request.form.get("notes", "")
        db.session.commit()
        flash("의뢰가 수정되었습니다.", "success")
        return redirect(url_for("request_detail", req_id=req_id))

    requesters = User.query.filter_by(role="requester").all()
    catalog = EfficacyCatalog.query.all()
    return render_template("request_edit.html", req=req, requesters=requesters, catalog=catalog)


@app.route("/requests/<int:req_id>/delete", methods=["POST"])
@login_required
def delete_request(req_id):
    if current_user.role != "admin":
        flash("관리자만 삭제할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req:
        flash("의뢰를 찾을 수 없습니다.", "error")
        return redirect(url_for("request_list"))

    Message.query.filter_by(request_id=req_id).delete()
    db.session.delete(req)
    db.session.commit()
    flash(f"의뢰 [{req.material_name}]이 삭제되었습니다.", "success")
    return redirect(url_for("request_list"))


# ── SharePoint URL ────────────────────────────────────
@app.route("/requests/<int:req_id>/set_sharepoint", methods=["POST"])
@login_required
def set_sharepoint(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.researcher_id != current_user.id or req.status != "documenting":
        flash("권한이 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req.sharepoint_url = request.form.get("sharepoint_url", "")
    db.session.commit()
    flash("SharePoint 링크가 등록되었습니다.", "success")
    # Redirect to parent if child
    redirect_id = req.parent_id if req.parent_id else req_id
    return redirect(url_for("request_detail", req_id=redirect_id))


# ── Researcher Claims ────────────────────────────────
@app.route("/requests/<int:req_id>/claim", methods=["POST"])
@login_required
def claim_request(req_id):
    if current_user.role != "researcher":
        flash("실험자만 접수할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req = db.session.get(Request, req_id)
    if not req or req.status != "pending":
        flash("접수할 수 없는 의뢰입니다.", "error")
        return redirect(url_for("dashboard"))

    # Specialized: only assigned researcher or team members can claim
    if req.is_specialized_child:
        allowed = False
        if req.assigned_researcher_id and req.assigned_researcher_id == current_user.id:
            allowed = True
        elif req.assigned_team and req.assigned_team == current_user.team:
            allowed = True
        elif not req.assigned_researcher_id and not req.assigned_team:
            allowed = False  # not yet assigned
        if not allowed:
            flash("이 전문평가는 배정된 실험자/팀만 접수할 수 있습니다.", "error")
            return redirect(url_for("dashboard"))

    req.researcher_id = current_user.id
    req.status = "in_progress"
    req.start_week = request.form.get("start_week", "")
    req.claimed_at = datetime.now(timezone.utc)
    db.session.commit()
    if req.parent_id:
        sync_parent_status(req.parent)
    actual_requester_id = req.parent.requester_id if req.parent_id else req.requester_id
    create_notification(actual_requester_id, f"[{req.material_name}] 실험자 {current_user.name}님이 접수했습니다.", f"/requests/{req.parent_id or req.id}")
    flash(f"의뢰 [{req.material_name}]을 접수했습니다.", "success")
    return redirect(url_for("request_detail", req_id=req_id))


# ── Status Transitions ───────────────────────────────
@app.route("/requests/<int:req_id>/to_documenting", methods=["POST"])
@login_required
def to_documenting(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.researcher_id != current_user.id or req.status != "in_progress":
        flash("상태를 변경할 수 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req.status = "documenting"
    db.session.commit()
    if req.parent_id:
        sync_parent_status(req.parent)
    flash("자료정리 단계로 전환되었습니다. SharePoint 링크를 등록하고 완료하세요.", "info")
    return redirect(url_for("request_detail", req_id=req_id))


@app.route("/requests/<int:req_id>/complete", methods=["POST"])
@login_required
def complete_request(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.researcher_id != current_user.id or req.status != "documenting":
        flash("상태를 변경할 수 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    result_text = request.form.get("result", "").strip()
    if not result_text:
        flash("종합 의견을 입력해주세요.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req.result = result_text
    req.status = "completed"
    req.gmail_sent = True
    req.completed_at = datetime.now(timezone.utc)

    # Handle PPT file upload
    file = request.files.get("ppt_file")
    if file and file.filename:
        fname = secure_filename(file.filename)
        unique_name = f"{req_id}_{int(datetime.now().timestamp())}_{fname}"
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
        file.save(save_path)
        req.file_path = unique_name
        req.file_name = fname
        req.file_uploaded_at = datetime.now(timezone.utc)

    db.session.commit()
    if req.parent_id:
        sync_parent_status(req.parent)
    # Notify requester
    actual_requester_id = req.parent.requester_id if req.parent_id else req.requester_id
    eff_label = f" - {req.efficacy_types}" if req.parent_id else ""
    create_notification(actual_requester_id, f"[{req.material_name}{eff_label}] 평가가 완료되었습니다.", f"/requests/{req.parent_id or req.id}")
    flash("평가가 완료되었습니다. 결과가 연간 리스트에 추가되었습니다.", "success")
    redirect_id = req.parent_id if req.parent_id else req_id
    return redirect(url_for("request_detail", req_id=redirect_id))


# ── File Download ────────────────────────────────────
@app.route("/requests/<int:req_id>/download_file")
@login_required
def download_file(req_id):
    req = db.session.get(Request, req_id)
    if not req or not req.file_path:
        flash("파일이 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    # Check expiry (7 days)
    if req.file_uploaded_at:
        expiry = req.file_uploaded_at + timedelta(days=7)
        if datetime.now(timezone.utc) > expiry:
            # Delete expired file
            fpath = os.path.join(app.config["UPLOAD_FOLDER"], req.file_path)
            if os.path.exists(fpath):
                os.remove(fpath)
            req.file_path = ""
            req.file_name = ""
            req.file_uploaded_at = None
            db.session.commit()
            flash("파일 다운로드 기간(1주일)이 만료되었습니다.", "error")
            return redirect(url_for("request_detail", req_id=req_id))

    fpath = os.path.join(app.config["UPLOAD_FOLDER"], req.file_path)
    if not os.path.exists(fpath):
        flash("파일을 찾을 수 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    return send_file(fpath, as_attachment=True, download_name=req.file_name)


# ── File Re-request ──────────────────────────────────
@app.route("/requests/<int:req_id>/request_file", methods=["POST"])
@login_required
def request_file(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.requester_id != current_user.id:
        flash("권한이 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    req.file_requested = True
    db.session.commit()
    flash("실험자에게 결과 파일을 재요청했습니다.", "success")
    return redirect(url_for("request_detail", req_id=req_id))


@app.route("/requests/<int:req_id>/reupload_file", methods=["POST"])
@login_required
def reupload_file(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.researcher_id != current_user.id:
        flash("권한이 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    file = request.files.get("ppt_file")
    if file and file.filename:
        fname = secure_filename(file.filename)
        unique_name = f"{req_id}_{int(datetime.now().timestamp())}_{fname}"
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
        file.save(save_path)
        req.file_path = unique_name
        req.file_name = fname
        req.file_uploaded_at = datetime.now(timezone.utc)
        req.file_requested = False
        db.session.commit()
        flash("파일을 재업로드했습니다.", "success")
    else:
        flash("파일을 선택해주세요.", "error")

    return redirect(url_for("request_detail", req_id=req_id))


# ── Chat ─────────────────────────────────────────────
@app.route("/requests/<int:req_id>/messages", methods=["POST"])
@login_required
def send_message(req_id):
    req = db.session.get(Request, req_id)
    if not req:
        return jsonify({"error": "의뢰를 찾을 수 없습니다."}), 400

    # Allow chat: in_progress (researcher-requester including child reqs)
    actual_requester_id = req.parent.requester_id if req.parent_id else req.requester_id
    if req.status == "in_progress":
        if current_user.id not in (actual_requester_id, req.researcher_id) and current_user.role != "admin":
            return jsonify({"error": "권한이 없습니다."}), 403
    elif req.status == "submitted" and req.has_specialized:
        if current_user.id != req.requester_id and current_user.role != "admin":
            return jsonify({"error": "권한이 없습니다."}), 403
    else:
        return jsonify({"error": "채팅이 비활성화되었습니다."}), 400

    content = request.form.get("content", "").strip()
    if not content:
        return jsonify({"error": "메시지를 입력하세요."}), 400

    msg = Message(request_id=req_id, sender_id=current_user.id, content=content)
    db.session.add(msg)
    db.session.commit()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({
            "id": msg.id,
            "sender": current_user.name,
            "content": msg.content,
            "time": msg.created_at.strftime("%m/%d %H:%M"),
        })

    return redirect(url_for("request_detail", req_id=req_id))


@app.route("/requests/<int:req_id>/messages/poll")
@login_required
def poll_messages(req_id):
    after_id = request.args.get("after", 0, type=int)
    msgs = Message.query.filter(
        Message.request_id == req_id, Message.id > after_id
    ).order_by(Message.created_at.asc()).all()

    return jsonify([{
        "id": m.id,
        "sender": m.sender.name,
        "sender_id": m.sender_id,
        "content": m.content,
        "time": m.created_at.strftime("%m/%d %H:%M"),
    } for m in msgs])


# ── Yearly Results ───────────────────────────────────
@app.route("/results")
@login_required
def yearly_results():
    year_filter = request.args.get("year", datetime.now().year, type=int)

    # Get completed parent/standalone requests (no children duplicates)
    completed_parents = Request.query.filter(
        Request.parent_id.is_(None),
        Request.status == "completed",
        Request.year == year_filter
    ).order_by(Request.completed_at.desc()).all()

    # Also get parents whose ALL children are completed
    completed_via_children = Request.query.filter(
        Request.parent_id.is_(None),
        Request.status.in_(["completed", "approved_parent"]),
        Request.year == year_filter,
        ~Request.children.any(Request.status != "completed")
    ).filter(Request.children.any()).order_by(Request.completed_at.desc()).all()

    # Merge and deduplicate
    seen_ids = set()
    results = []
    for r in completed_parents + completed_via_children:
        if r.id not in seen_ids:
            seen_ids.add(r.id)
            results.append(r)

    years = db.session.query(db.func.distinct(Request.year)).filter(
        Request.status == "completed").order_by(Request.year.desc()).all()
    years = [y[0] for y in years if y[0]]
    if year_filter not in years and years:
        year_filter = years[0]

    # Stats from child requests
    all_completed_children = Request.query.filter(
        Request.status == "completed", Request.year == year_filter
    ).all()
    total = len(results)
    by_team = {}
    by_efficacy = {}
    for r in all_completed_children:
        team = r.requester_team or (r.parent.requester_team if r.parent else "")
        if team:
            by_team[team] = by_team.get(team, 0) + 1
        for eff in r.efficacy_types.split(","):
            eff = eff.strip()
            if eff:
                by_efficacy[eff] = by_efficacy.get(eff, 0) + 1

    # Researcher ranking
    by_researcher = {}
    for r in all_completed_children:
        if r.researcher:
            key = r.researcher.name
            if key not in by_researcher:
                by_researcher[key] = {"name": r.researcher.name, "team": r.researcher.team, "count": 0}
            by_researcher[key]["count"] += 1
    researcher_ranking = sorted(by_researcher.values(), key=lambda x: x["count"], reverse=True)

    # Requester ranking (by team and individual)
    by_requester_team = {}
    by_requester = {}
    for r in all_completed_children:
        req_team = r.requester_team or (r.parent.requester_team if r.parent else "")
        requester = r.parent.requester if r.parent else r.requester
        if req_team:
            by_requester_team[req_team] = by_requester_team.get(req_team, 0) + 1
        if requester:
            key = requester.name
            if key not in by_requester:
                by_requester[key] = {"name": requester.name, "team": requester.team, "count": 0}
            by_requester[key]["count"] += 1
    requester_team_ranking = sorted(by_requester_team.items(), key=lambda x: x[1], reverse=True)
    requester_ranking = sorted(by_requester.values(), key=lambda x: x["count"], reverse=True)

    return render_template("yearly_results.html", results=results, years=years,
                           year_filter=year_filter, total=total, by_team=by_team,
                           by_efficacy=by_efficacy, researcher_ranking=researcher_ranking,
                           requester_team_ranking=requester_team_ranking,
                           requester_ranking=requester_ranking)


# ── Excel Download ───────────────────────────────────
@app.route("/results/download")
@login_required
def download_results():
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year_filter = request.args.get("year", datetime.now().year, type=int)
    results = Request.query.filter_by(status="completed", year=year_filter).order_by(Request.completed_at.desc()).all()

    # Gather stats
    by_team = {}
    by_efficacy = {}
    by_researcher = {}
    for r in results:
        by_team[r.requester_team] = by_team.get(r.requester_team, 0) + 1
        for eff in r.efficacy_types.split(","):
            eff = eff.strip()
            if eff:
                by_efficacy[eff] = by_efficacy.get(eff, 0) + 1
        if r.researcher:
            key = r.researcher.name
            if key not in by_researcher:
                by_researcher[key] = {"name": r.researcher.name, "team": r.researcher.team, "count": 0}
            by_researcher[key]["count"] += 1
    researcher_ranking = sorted(by_researcher.values(), key=lambda x: x["count"], reverse=True)

    wb = Workbook()

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    sub_header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    def style_header(ws, row, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def style_data_cell(cell, align_center=False):
        cell.border = thin_border
        cell.alignment = center if align_center else wrap

    # ══════════════════════════════════════════
    # Sheet 1: 의뢰 평가 결과 목록
    # ══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = f"{year_filter}년 평가결과"

    headers = ["번호", "소재", "의뢰팀", "의뢰자", "효능", "농도", "샘플수",
               "대조군", "대조군명", "대조군농도", "긴급도", "마감",
               "실험자", "실험자팀", "예상(주)", "상태", "결과 요약",
               "등록일", "접수일", "완료일"]

    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers))

    for i, r in enumerate(results, 2):
        row_data = [
            r.id, r.material_name, r.requester_team, r.requester.name,
            r.efficacy_types, r.concentration, r.sample_count,
            "O" if r.has_control else "", r.control_name, r.control_concentration,
            r.urgency, r.deadline,
            r.researcher.name if r.researcher else "", r.researcher.team if r.researcher else "",
            r.estimated_weeks or "", "완료", r.result,
            r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
            r.claimed_at.strftime("%Y-%m-%d") if r.claimed_at else "",
            r.completed_at.strftime("%Y-%m-%d") if r.completed_at else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            style_data_cell(cell, align_center=(col not in (2, 5, 9, 17)))
        # Alternate row color
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws1.cell(row=i, column=col).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Column widths
    col_widths = [6, 20, 8, 8, 30, 10, 7, 7, 15, 10, 8, 12, 8, 8, 8, 7, 40, 11, 11, 11]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.auto_filter.ref = f"A1:T{len(results) + 1}"
    ws1.freeze_panes = "A2"

    # ══════════════════════════════════════════
    # Sheet 2: 팀별 의뢰 통계 + 파이 차트
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet(title="팀별 의뢰 통계")

    ws2.cell(row=1, column=1, value="팀")
    ws2.cell(row=1, column=2, value="건수")
    ws2.cell(row=1, column=3, value="비율(%)")
    style_header(ws2, 1, 3)

    total = len(results) or 1
    sorted_teams = sorted(by_team.items(), key=lambda x: x[1], reverse=True)
    for i, (team, count) in enumerate(sorted_teams, 2):
        ws2.cell(row=i, column=1, value=team)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=round(count / total * 100, 1))
        for col in range(1, 4):
            style_data_cell(ws2.cell(row=i, column=col), align_center=True)

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    if sorted_teams:
        pie1 = PieChart()
        pie1.title = f"{year_filter}년 팀별 의뢰 비율"
        pie1.style = 10
        pie1.width = 18
        pie1.height = 14
        data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(sorted_teams) + 1)
        labels_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(sorted_teams) + 1)
        pie1.add_data(data_ref, titles_from_data=True)
        pie1.set_categories(labels_ref)
        ws2.add_chart(pie1, "E2")

    # ══════════════════════════════════════════
    # Sheet 3: 효능별 건수 통계 + 파이 차트
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet(title="효능별 통계")

    ws3.cell(row=1, column=1, value="효능")
    ws3.cell(row=1, column=2, value="건수")
    ws3.cell(row=1, column=3, value="비율(%)")
    style_header(ws3, 1, 3)

    total_eff = sum(by_efficacy.values()) or 1
    sorted_effs = sorted(by_efficacy.items(), key=lambda x: x[1], reverse=True)
    for i, (eff, count) in enumerate(sorted_effs, 2):
        ws3.cell(row=i, column=1, value=eff)
        ws3.cell(row=i, column=2, value=count)
        ws3.cell(row=i, column=3, value=round(count / total_eff * 100, 1))
        for col in range(1, 4):
            style_data_cell(ws3.cell(row=i, column=col), align_center=True)

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 10

    if sorted_effs:
        pie2 = PieChart()
        pie2.title = f"{year_filter}년 효능별 평가 비율"
        pie2.style = 10
        pie2.width = 18
        pie2.height = 14
        data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(sorted_effs) + 1)
        labels_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(sorted_effs) + 1)
        pie2.add_data(data_ref, titles_from_data=True)
        pie2.set_categories(labels_ref)
        ws3.add_chart(pie2, "E2")

    # ══════════════════════════════════════════
    # Sheet 4: 실험자별 랭킹
    # ══════════════════════════════════════════
    ws4 = wb.create_sheet(title="실험자 랭킹")

    ws4.cell(row=1, column=1, value="순위")
    ws4.cell(row=1, column=2, value="실험자")
    ws4.cell(row=1, column=3, value="소속팀")
    ws4.cell(row=1, column=4, value="완료 건수")
    style_header(ws4, 1, 4)

    for i, r in enumerate(researcher_ranking, 2):
        ws4.cell(row=i, column=1, value=i - 1)
        ws4.cell(row=i, column=2, value=r["name"])
        ws4.cell(row=i, column=3, value=r["team"])
        ws4.cell(row=i, column=4, value=r["count"])
        for col in range(1, 5):
            style_data_cell(ws4.cell(row=i, column=col), align_center=True)
        # Gold / Silver / Bronze
        rank_colors = {1: "FEF3C7", 2: "F3F4F6", 3: "FFEDD5"}
        if (i - 1) in rank_colors:
            for col in range(1, 5):
                ws4.cell(row=i, column=col).fill = PatternFill(
                    start_color=rank_colors[i - 1], end_color=rank_colors[i - 1], fill_type="solid")

    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 12

    # ── Save & Return ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BT_Lab_효능평가결과_{year_filter}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ── Helper: create notification ─────────────────────
def create_notification(user_id, content, link=""):
    notif = Notification(user_id=user_id, content=content, link=link)
    db.session.add(notif)
    db.session.commit()


# ── Raw Data Request (per-efficacy / per-child) ─────
@app.route("/requests/<int:req_id>/raw_data_request", methods=["POST"])
@login_required
def raw_data_request(req_id):
    req = db.session.get(Request, req_id)
    if not req or req.status != "completed":
        flash("완료된 의뢰에만 raw data를 요청할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    # Only the requester of this request (or parent request) can request
    actual_requester_id = req.parent.requester_id if req.parent_id else req.requester_id
    if current_user.id != actual_requester_id:
        flash("의뢰자만 raw data를 요청할 수 있습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    # This route now handles individual requests (child or standalone)
    if not req.researcher_id:
        flash("담당 실험자가 없습니다.", "error")
        return redirect(url_for("request_detail", req_id=req_id))

    existing = RawDataRequest.query.filter_by(
        request_id=req.id, requester_id=current_user.id,
        researcher_id=req.researcher_id, status="requested"
    ).first()
    if existing:
        flash("이미 raw data 요청을 보냈습니다.", "info")
    else:
        rdr = RawDataRequest(
            request_id=req.id, requester_id=current_user.id,
            researcher_id=req.researcher_id
        )
        db.session.add(rdr)
        material = req.parent.material_name if req.parent_id else req.material_name
        eff_label = f" - {req.efficacy_types}" if req.parent_id else ""
        create_notification(
            req.researcher_id,
            f"[{material}{eff_label}] raw data 요청이 도착했습니다.",
            f"/requests/{req.id}"
        )
        db.session.commit()
        flash("실험자에게 raw data 요청을 보냈습니다.", "success")

    redirect_id = req.parent_id if req.parent_id else req_id
    return redirect(url_for("request_detail", req_id=redirect_id))


@app.route("/raw_data/<int:rdr_id>/send", methods=["POST"])
@login_required
def send_raw_data(rdr_id):
    rdr = db.session.get(RawDataRequest, rdr_id)
    if not rdr or rdr.researcher_id != current_user.id:
        flash("권한이 없습니다.", "error")
        return redirect(url_for("dashboard"))

    file = request.files.get("raw_data_file")
    if not file or not file.filename:
        flash("파일을 선택해주세요.", "error")
        return redirect(url_for("request_detail", req_id=rdr.request_id))

    fname = secure_filename(file.filename)
    unique_name = f"rawdata_{rdr.id}_{int(datetime.now().timestamp())}_{fname}"
    save_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
    file.save(save_path)

    rdr.file_path = unique_name
    rdr.file_name = fname
    rdr.status = "sent"
    rdr.sent_at = datetime.now(timezone.utc)
    db.session.commit()

    # Notify requester
    req = db.session.get(Request, rdr.request_id)
    material = req.material_name if req else ""
    create_notification(
        rdr.requester_id,
        f"[{material}] raw data 파일이 전송되었습니다.",
        f"/requests/{rdr.request_id}"
    )

    flash("raw data 파일을 전송했습니다.", "success")
    redirect_id = req.parent_id if req and req.parent_id else rdr.request_id
    return redirect(url_for("request_detail", req_id=redirect_id))


@app.route("/raw_data/<int:rdr_id>/download")
@login_required
def download_raw_data(rdr_id):
    rdr = db.session.get(RawDataRequest, rdr_id)
    if not rdr or not rdr.file_path:
        flash("파일이 없습니다.", "error")
        return redirect(url_for("dashboard"))

    if current_user.id not in (rdr.requester_id, rdr.researcher_id) and current_user.role != "admin":
        flash("권한이 없습니다.", "error")
        return redirect(url_for("dashboard"))

    fpath = os.path.join(app.config["UPLOAD_FOLDER"], rdr.file_path)
    if not os.path.exists(fpath):
        flash("파일을 찾을 수 없습니다.", "error")
        return redirect(url_for("dashboard"))

    return send_file(fpath, as_attachment=True, download_name=rdr.file_name)


# ── Notifications ────────────────────────────────────
@app.route("/notifications")
@login_required
def notifications_page():
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(
        Notification.created_at.desc()).all()
    # Mark all as read
    for n in notifs:
        if not n.is_read:
            n.is_read = True
    db.session.commit()
    return render_template("notifications.html", notifications=notifs)


@app.route("/notifications/unread_count")
@login_required
def unread_notification_count():
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({"count": count})


# ── Init DB & Seed ───────────────────────────────────
def seed_data():
    if User.query.first():
        return

    # Admin
    db.session.add(User(username="admin", password=generate_password_hash("admin123"),
                        name="관리자", team="BT Lab", role="admin", email="admin@btlab.com"))
    db.session.add(User(username="nyh", password=generate_password_hash("1234"),
                        name="노윤화", team="BT Lab", role="admin", email="nyh@btlab.com"))

    # Researchers (BT1, BT2, BT3)
    researchers = [
        ("lke", "이경은", "BT1"), ("kmj", "김민지", "BT1"), ("pjy", "박정연", "BT1"),
        ("pej", "박은진", "BT1"), ("jhy", "조희연", "BT1"), ("rdy", "류다연", "BT1"),
        ("kms", "김미선", "BT2"), ("ks", "김솔", "BT2"), ("khy", "김혜연", "BT2"),
        ("kmsg", "김민상", "BT2"), ("pyj", "박연지", "BT2"), ("ksy", "경서연", "BT1"),
        ("ksh", "김세현", "BT3"), ("rkm", "류경민", "BT3"), ("lsh", "임소희", "BT3"),
    ]
    for uname, name, team in researchers:
        db.session.add(User(username=uname, password=generate_password_hash("1234"),
                            name=name, team=team, role="researcher", email=f"{uname}@btlab.com"))

    # Requesters
    requesters = [
        ("lhe", "이하은", "MB2"), ("kjh", "김지현", "BI1"), ("cyj", "최유진", "BI2"),
        ("phj", "박현준", "MB1"), ("sjw", "송지원", "BI3"),
    ]
    for uname, name, team in requesters:
        db.session.add(User(username=uname, password=generate_password_hash("1234"),
                            name=name, team=team, role="requester", email=f"{uname}@btlab.com"))

    # Efficacy Catalog
    common = [
        ("항노화", "MMP-1", "HS68 (fibroblast)"),
        ("탄력", "Col, ELN, FBN", "HS68 (fibroblast)"),
        ("재생", "Migration", "HS68 (fibroblast)"),
        ("보습/수분", "HAS3, AQP3", "HaCaT (keratinocyte)"),
        ("장벽", "FLG, CLDN, IVL", "HaCaT (keratinocyte)"),
        ("진정", "IL-1b, IL-6, TNFa", "HaCaT (keratinocyte)"),
        ("가려움 개선", "TSLP", "HaCaT (keratinocyte)"),
        ("표피 증식", "PCNA, KI67", "HaCaT (keratinocyte)"),
        ("멜라닌 억제", "Melanin contents", "B16F10 (Melanocyte)"),
        ("지질 억제", "SREBP", "SZ95 (Sebocyte)"),
        ("독성", "Cell viability", "HS68, HaCaT, B16F10"),
        ("항산화", "-", "DPPH solution"),
        ("냉감", "TRPM8, CIRBP", "HaCaT (keratinocyte)"),
    ]
    for name, marker, cell in common:
        db.session.add(EfficacyCatalog(name=name, target_marker=marker, cell_line=cell,
                                       bt_group="common", requires_discussion=False))

    bt1_special = ["열노화", "립", "저산소", "미세먼지", "반려동물"]
    for name in bt1_special:
        db.session.add(EfficacyCatalog(name=name, bt_group="BT1", requires_discussion=True))

    bt2_special = ["항노화", "3D skin", "explant", "흡수도"]
    for name in bt2_special:
        db.session.add(EfficacyCatalog(name=name, bt_group="BT2", requires_discussion=True))

    bt3_special = ["헤어", "두피 개선", "ATP"]
    for name in bt3_special:
        db.session.add(EfficacyCatalog(name=name, bt_group="BT3", requires_discussion=True))

    db.session.commit()
    print("Seed data created!")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        seed_data()
    app.run(debug=True, port=5000)
